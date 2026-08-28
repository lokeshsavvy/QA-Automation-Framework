from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import openpyxl
from OrangeHRM_Project.pages.loginpage import LoginPage


def test_login(driver):

    workbook = openpyxl.load_workbook("OrangeHRM_Project/testdata/login_data.xlsx")
    sheet = workbook.active

    for row in range(2,sheet.max_row+1):
        username = sheet.cell(row = row,column=6).value
        password = sheet.cell(row=row,column=7).value
        sheet.cell(row=row,column=5).value=(f"Login Test With Username={username}")
        login_page =LoginPage(driver)
        login_page.open_url("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='username']")))

        print(f"Executing Row: {row}")
        print("username:", username)
        print("password:", password)

        login_page.login(username,password)

        current_date = datetime.now().strftime("%d-%m-%y")
        sheet.cell(row=row, column=4).value = current_date

        tester_name =sheet.cell(row=row,column=3).value

        if 'dashboard' in driver.current_url.lower():
            sheet.cell(row=row,column=8).value = "Test Passed"
            login_page.logout()
        else:
            sheet.cell(row=row,column=8).value = "Test Failed"

    workbook.save("OrangeHRM_Project/testdata/login_data.xlsx")

def test_verify_login_fields(driver):

    login_page = LoginPage(driver)
    login_page.open_url("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

    assert login_page.is_username_displayed(),"user name field not displayed"
    assert login_page.is_password_displayed(),"password field is not displayed"
    assert login_page.is_username_enabled(),"user name field not enabled"
    assert login_page.is_password_enabled(),"password field is not enabled"

def test_forgot_password(driver):

    login = LoginPage(driver)
    login.open_url("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

    login.click_forgot_passwrod()
    login.enter_reset_username("Rahul99999")
    login.click_reset_password()

    assert login.is_reset_sucessful()






